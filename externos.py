import openpyxl
import re

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Externos Santander"

# Encabezados
ws["A1"] = "Nombre Completo"
ws["B1"] = "Empresa"

with open("externos_santander.txt", "r", encoding="utf-8") as f:
    contenido = f.read()

empresa_actual = ""
fila = 2

for linea in contenido.splitlines():
    linea = linea.strip()

    # Detectar empresa
    match_empresa = re.match(r"^===\s*(.+?)\s*===$", linea)
    if match_empresa:
        empresa_actual = match_empresa.group(1)

    # Detectar nombre
    elif linea.startswith("- "):
        nombre_completo = linea[2:].strip()
        ws[f"A{fila}"] = nombre_completo
        ws[f"B{fila}"] = empresa_actual
        fila += 1

wb.save("externos_santander.xlsx")
print(f"Archivo guardado con {fila - 2} personas.")